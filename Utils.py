import re
import csv
from openpyxl import load_workbook

def readxlsx (filepath):
    tempoMap = {}
    wb=load_workbook(filepath)
    sheet=wb.active
    max_row=sheet.max_row
    #max_column=sheet.max_column
    for i in range(2,max_row+1):
        key = None
        for j in range(4,6):
            cell_obj=sheet.cell(row=i,column=j)
            if (j==4):
                key = cell_obj.value
            if (j==5):
                tempoMap[key] = cell_obj.value
    return tempoMap

# Handle cf[11001] = 12
def updateStringPatternOne (content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *= *[0-9]+', v)
        for match in matches:
            num = match.split('=')[-1]
            v = re.sub(r'cf\[11001\] *= *[0-9]+', 'Team = \'' + str(tempoMap[int(num)]) +'\'', v, 1)
            content[k]=v

# Handle cf[11001] != 12
def updateStringPatternTwo (content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *!= *[0-9]+', v)
        for match in matches:
            num = match.split('=')[-1]
            v = re.sub(r'cf\[11001\] *!= *[0-9]+', 'Team != \'' + str(tempoMap[int(num)]) +'\'', v, 1)
            content[k]=v

# Handle cf[11001] in (12,13)
def updateStringPatternThree (content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *in *\([0-9, ]*\)',v)
        for match in matches:
            teamString = transformString(match.split('in')[-1],tempoMap)
            v = re.sub(r'cf\[11001\] *in *\([0-9, ]*\)', 'Team in ' + teamString, v, 1)
            content[k]=v

# Handle cf[11001] not in (12,13)
def updateStringPatternFour (content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *not *in *\([0-9, ]*\)',v)
        for match in matches:
            teamString = transformString(match.split('in')[-1],tempoMap)
            v = re.sub(r'cf\[11001\] *not *in *\([0-9, ]*\)', 'Team not in ' + teamString, v, 1)
            content[k]=v

# Helper method
def transformString(numberString,tempoMap):
    stringList = numberString.replace(' ','').replace('(','').replace(')','').split(',')
    newList=[]
    for number in stringList:
        newList.append(str(tempoMap[int(number)]))
    retString = '(\'' + '\',\''.join(newList) + '\')'
    return str(retString)

# Handle cf[11001] = "12"
def updateStringPatternFive(content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *= *"[0-9]+"', v)
        for match in matches:
            num = match.replace('\"','').split('=')[-1]
            v = re.sub(r'cf\[11001\] *= *"[0-9]+"', 'Team = \'' + str(tempoMap[int(num)]) +'\'', v, 1)
            content[k]=v

# Handle cf[11001] != "12"
def updateStringPatternSix(content,tempoMap):
    for k,v in content.items():
        matches = re.findall(r'cf\[11001\] *!= *"[0-9]+"', v)
        for match in matches:
            num = match.replace('\"','').split('=')[-1]
            v = re.sub(r'cf\[11001\] *!= *"[0-9]+"', 'Team != \'' + str(tempoMap[int(num)]) +'\'', v, 1)
            content[k]=v

# Replace remaining cf[11001] with Team
def replaceCf(content):
    for k,v in content.items():
        v = re.sub(r'cf\[11001\]', 'Team', v)
        content[k]=v

# Write Map to a csv file
def writeCsv(content):
    w = csv.writer(open('filter.csv','w'))
    w.writerow(['id','reqcontent'])
    for k,v in content.items():
        w.writerow([k,v])
